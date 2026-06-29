#!/usr/bin/env python3
"""Generate spanish_review.xlsx for Hongye Furniture bilingual review."""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Paths
BASE = os.path.dirname(os.path.abspath(__file__))
LOCALES = os.path.join(os.path.dirname(BASE), 'i18n', 'locales')
EN_PATH = os.path.join(LOCALES, 'en.json')
ES_PATH = os.path.join(LOCALES, 'es.json')
OUT_PATH = os.path.join(BASE, 'spanish_review.xlsx')

# Brand colours
BRAND_DARK = '121A22'
BRAND_WHITE = 'FFFFFF'
ALT_ROW = 'F2F4F6'
BORDER_COLOR = 'CBD0D6'

def flatten(obj, parent_key='', sep='.'):
    """Flatten nested dict into dot-notation keys, returning list of (key, value)."""
    items = []
    for k, v in obj.items():
        new_key = f'{parent_key}{sep}{k}' if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten(v, new_key, sep))
        elif isinstance(v, list):
            # Store lists as JSON string
            items.append((new_key, json.dumps(v, ensure_ascii=False)))
        else:
            items.append((new_key, str(v) if not isinstance(v, str) else v))
    return items

def extract_page_section(key):
    """Extract page and section from a dot-notation key."""
    parts = key.split('.')
    page_map = {
        'common': ('Common', 'Navigation / Footer'),
        'index': ('Index', ''),
        'about': ('About', ''),
        'products': ('Products', ''),
        'contact': ('Contact', ''),
        'page': ('Common', 'Page Meta'),
    }

    page = page_map.get(parts[0], (parts[0].capitalize(), ''))
    page_name = page[0]

    # Derive section from the second/third level of the key
    section = ''
    if parts[0] == 'common':
        if len(parts) > 1:
            section = parts[1].capitalize()
    elif parts[0] in ('index', 'about', 'products', 'contact'):
        if len(parts) > 1:
            section = parts[1].capitalize()

    return page_name, section


def main():
    print('Loading translations...')

    with open(EN_PATH, 'r', encoding='utf-8') as f:
        en_data = json.load(f)
    with open(ES_PATH, 'r', encoding='utf-8') as f:
        es_data = json.load(f)

    # Flatten both
    en_flat = dict(flatten(en_data))
    es_flat = dict(flatten(es_data))

    # Collect all keys (sorted) — skip array-only keys, keep them for reference
    all_keys = sorted(set(en_flat.keys()) | set(es_flat.keys()))

    # Build rows
    rows = []
    for key in all_keys:
        page, section = extract_page_section(key)
        en_val = en_flat.get(key, '')
        es_val = es_flat.get(key, '')
        rows.append((page, section, key, en_val, es_val, '', '', ''))

    print(f'  {len(rows)} translation entries')

    # -- Create workbook --
    wb = Workbook()
    ws = wb.active
    ws.title = 'Spanish Review'

    # Headers
    headers = ['Page', 'Section', 'Key', 'English (EN)', 'AI Spanish (ES)',
               'Status', 'Notes', 'Corrected Spanish']

    # Styles
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=BRAND_DARK, end_color=BRAND_DARK, fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    body_font = Font(name='Segoe UI', size=10)
    body_align = Alignment(vertical='top', wrap_text=True)
    alt_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR),
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border
            # Alternating row colour
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    col_widths = [14, 18, 36, 54, 54, 14, 30, 54]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Auto-filter
    last_row = len(rows) + 1
    ws.auto_filter.ref = f'A1:H{last_row}'

    # Freeze header row
    ws.freeze_panes = 'A2'

    # -- Data Validation: Status dropdown --
    status_dv = DataValidation(
        type='list',
        formula1='"待审核,需修改,已通过"',
        allow_blank=True,
    )
    status_dv.error = 'Please select from the dropdown.'
    status_dv.errorTitle = 'Invalid Status'
    ws.add_data_validation(status_dv)
    status_dv.add(f'F2:F{last_row}')

    # Save
    wb.save(OUT_PATH)
    print(f'\nDone! Review file saved to:\n  {OUT_PATH}')


if __name__ == '__main__':
    main()
